import os
import sys
from io import StringIO
import traceback
from pathlib import Path
import elevate
from datetime import datetime
import kivy
from kivy.config import Config
from kivy.core.window import Window
from kivy.app import App
from kivy.uix.dropdown import DropDown
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.switch import Switch
from openpyxl import Workbook
from openpyxl import load_workbook
from kivy.uix.popup import Popup

elevate.elevate()

class FolderCreator(App):
    def build(self):     
        Window.size = (600,380)

        self.kwtmva = ""
        self.ueber_ordner = ""
           
        layout = GridLayout(cols= 2, spacing=5, padding= 30)
        widget_height = 60

        geber_label = Label(text= "Auftraggeber: ",size_hint_y=None, height=30)
        layout.add_widget(geber_label)

        self.geber_text = TextInput(multiline= False,size_hint_y=None, height=30, write_tab= False)
        layout.add_widget(self.geber_text)

        arbeit_label = Label(text= "Projekt: ",size_hint_y=None, height=30)
        layout.add_widget(arbeit_label)

        self.arbeit_text = TextInput(multiline= False,size_hint_y=None, height=30, write_tab= False)
        layout.add_widget(self.arbeit_text)

        strasse_label = Label(text= "Straße: ",size_hint_y=None, height=30)
        layout.add_widget(strasse_label)

        self.strasse_text = TextInput(multiline= False,size_hint_y=None, height=30, write_tab= False)
        layout.add_widget(self.strasse_text)

        plz_label = Label(text= "PLZ / Ort: ",size_hint_y=None, height=30)
        layout.add_widget(plz_label)

        self.plz_text = TextInput(multiline= False,size_hint_y=None, height=30, write_tab= False)
        layout.add_widget(self.plz_text)

        telefon_label = Label(text= "Telefonnr.: ",size_hint_y=None, height=30)
        layout.add_widget(telefon_label)

        self.telefon_text = TextInput(multiline= False,size_hint_y=None, height=30, write_tab= False)
        layout.add_widget(self.telefon_text)

        summe_label = Label(text= "Auftragssumme: ",size_hint_y=None, height=30)
        layout.add_widget(summe_label)

        self.summe_text = TextInput(multiline= False,size_hint_y=None, height=30, write_tab= False)
        layout.add_widget(self.summe_text)

        self.auswahl_label = Label(text= f"Aktuelle Auftragsnummer KAU: {self.get_current_kau_number()}")
        layout.add_widget(self.auswahl_label)    
        
        self.button_label = Label(text= f"Aktuelle Auftragsnummer MVA/KWT: {self.get_current_number()}")
        layout.add_widget(self.button_label)


        self.dropdown_button = Button(text="Ordner auswählen",size_hint_y=None, height=widget_height)
        self.dropdown_button.bind(on_release=self.show_dropdown)

        self.dropdown = DropDown()
        for item in ['Kleinstaufträge', 'Klärwerkstechnik', 'Metallverarbeitung']:
            btn = Button(text=item, size_hint_y=None, height=widget_height)
            btn.bind(on_release=self.select_dropdown_option)
            self.dropdown.add_widget(btn)
        layout.add_widget(self.dropdown_button)

        erstellen_button = Button(text= "Ordner Erstellen", on_press=self.create_folder,size_hint_y=None, height=widget_height)
        layout.add_widget(erstellen_button)

        return layout


    def show_dropdown(self, instance):
        self.dropdown.open(instance)


    def select_dropdown_option(self, instance):
        print(f"Ausgewählte Option: {instance.text}")
        self.dropdown_button.text = instance.text
        self.dropdown.dismiss()
        if instance.text == "Kleinstaufträge":
            self.ueber_ordner = "04-Kleinstaufträge"
            self.kwtmva = "KAU_"
            self.current_number = self.get_current_kau_number()
            self.number_text = "current_kau_number.txt"

        elif instance.text == "Klärwerkstechnik":
            self.ueber_ordner = "03-Aufträge"
            self.kwtmva = "KWT_"
            self.current_number = self.get_current_number()
            self.number_text = "current_number.txt"

        elif instance.text == "Metallverarbeitung":
            self.ueber_ordner = "03-Aufträge"
            self.kwtmva = "MVA_"
            self.current_number = self.get_current_number()
            self.number_text = "current_number.txt"


    def get_current_number(self):
        if os.path.exists("current_number.txt"):
            with open("current_number.txt", "r") as file:
                return int(file.read().strip())
        else:
            return 1


    def get_current_kau_number(self):
        if os.path.exists("current_kau_number.txt"):
            with open("current_kau_number.txt", "r") as file:
                return int(file.read().strip())
        else:
            return 1


    def create_folder(self,instance):

        if getattr(sys, 'frozen', False):
            working_dir = os.path.dirname(sys.executable)
        else:
            working_dir = os.path.dirname(__file__)

        os.chdir(working_dir)

        auftragsname = self.geber_text.text
        arbeit = self.arbeit_text.text
        datum = datetime.now().strftime("%Y%m%d")
        auftrag_ordner_name = f"{self.current_number:03d}_{datum}_{self.kwtmva}{auftragsname}-{arbeit}"
        texteingabe = f"Erstellt am: {datum} \nAuftraggeber: {auftragsname} \nProje: {arbeit} \nStraße: {self.strasse_text.text} \nPLZ / Ort: {self.plz_text.text} \nTelefonnr.: {self.telefon_text.text} \n"
        ueber_ordner = self.ueber_ordner        
        manufactory_ordner = "07-Manufactory"
        ueber_dir = os.path.join(os.getcwd(), ueber_ordner, auftrag_ordner_name)
        manufactory_dir = os.path.join(os.getcwd(), manufactory_ordner, auftrag_ordner_name)
        ordner_ver = os.path.join(working_dir, ueber_ordner)

        if ueber_ordner == "":
            error_message = ("Bitte wählen Sie einen Ordner aus.")
            self.show_error(error_message)
            return

        if not os.path.exists(ueber_ordner):
            os.makedirs(ueber_ordner)
            print(f"Der Ordner '{ueber_ordner}' wurde erfolgreich erstellt.")
        else:
            print(f"Der Ordner '{ueber_ordner}' existiert bereits.")

        if not os.path.exists(manufactory_ordner):
            os.makedirs(manufactory_ordner)
            print(f"Der Ordner '{manufactory_ordner}' wurde erfolgreich erstellt.")
        else:
            print(f"Der Ordner '{manufactory_ordner}' existiert bereits.")

        os.chdir(manufactory_ordner)

        if not os.path.exists(auftrag_ordner_name):
            os.makedirs(auftrag_ordner_name)
            print(f"Der Ordner '{auftrag_ordner_name}' wurde erfolgreich im {manufactory_ordner} erstellt.")
        else:
            print(f"Der Ordner '{auftrag_ordner_name}' in {manufactory_ordner} existiert bereits.")

        with open(os.path.join(auftrag_ordner_name, f"{auftragsname}_{arbeit}.txt"), "w") as file:
                file.write(texteingabe)

        try:
            os.chdir(ordner_ver)
            print(f"Das Ordnerverzeichnis '{ordner_ver}' wurde gefunden.")
        except FileNotFoundError:
            error_message = "Das Ordnerverzeichnis '{ordner_ver}' wurde nicht gefunden."
            self.show_error(error_message)

        if not os.path.exists(auftrag_ordner_name):      
            os.makedirs(auftrag_ordner_name)
            print(f"Der Ordner '{auftrag_ordner_name}' wurde erfolgreich erstellt.")

            try:
                manufactory = f"MFC_{auftrag_ordner_name}"
                manufactory_ver = os.path.join(ueber_dir, manufactory)
                os.symlink(manufactory_dir, manufactory_ver)
                print(f"Die Ordnerverknüpfung {manufactory} wurde erfolgreich erstellt.")
            except PermissionError:
                error_message = "keine Rechte vorhanden. Programm als Admin starten!"
                self.show_error(error_message)

            with open(os.path.join(auftrag_ordner_name, f"{auftragsname}_{arbeit}.txt"), "w") as file:
                file.write(texteingabe)

        else:
            print(f"Der Ordner '{auftrag_ordner_name}' existiert bereits.")
        
        os.chdir(working_dir)
        with open(self.number_text, "w") as file:
            file.write(str(self.current_number + 1))

        self.exelliste()


    def exelliste(self):
        dateiname = "Auftragsliste.xlsx"
        pfad = Path(dateiname)

        try:
            if pfad.is_file():
                wb = load_workbook(filename="Auftragsliste.xlsx")
                ws = wb.active
                row_number = ws.max_row + 1

                ws["A" + str(row_number)] = self.current_number
                ws["B" + str(row_number)] = datetime.now().strftime("%d.%m.%Y")
                ws["C" + str(row_number)] = self.kwtmva
                ws["D" + str(row_number)] = self.geber_text.text
                ws["E" + str(row_number)] = self.arbeit_text.text
                ws["F" + str(row_number)] = self.strasse_text.text
                ws["G" + str(row_number)] = self.plz_text.text
                ws["H" + str(row_number)] = self.telefon_text.text
                ws["I" + str(row_number)] = self.summe_text.text

                wb.save(dateiname)
                print(f"Die Daten wurden zur {dateiname} hinzugefügt.")

            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Auftragsliste"

                ws["A1"] = "Nr"
                ws["B1"] = "Datum"
                ws["C1"] = "KWT/MVA"
                ws["D1"] = "Auftraggeber"
                ws["E1"] = "Projekt"
                ws["F1"] = "Straße"
                ws["G1"] = "PLZ / Ort"
                ws["H1"] = "Telefonnr."
                ws["I1"] = "Summe"

                ws["A2"] = self.current_number
                ws["B2"] = datetime.now().strftime("%d.%m.%Y")
                ws["C2"] = self.kwtmva
                ws["D2"] = self.geber_text.text
                ws["E2"] = self.arbeit_text.text
                ws["F2"] = self.strasse_text.text
                ws["G2"] = self.plz_text.text
                ws["H2"] = self.telefon_text.text
                ws["I2"] = self.summe_text.text

                wb.save(dateiname)
                print(f"Die Exelliste {dateiname} existiert nicht. Eine neue Exeltabelle wird erstellt")

            self.clear_text()

        except PermissionError:
            error_message = "Auftragsliste ist geöffnet. Bitte schließen Sie die Datei."
            self.show_error(error_message)

            
    def clear_text(self):
        self.geber_text.text = ""
        self.arbeit_text.text = ""
        self.strasse_text.text = ""
        self.plz_text.text = ""
        self.telefon_text.text = ""
        self.summe_text.text = ""
        self.current_number += 1
        self.button_label.text = f"Aktuelle Auftragsnummer MVA/KWT: {self.get_current_number()}"
        self.auswahl_label.text = f"Aktuelle Auftragsnummer KAU: {self.get_current_kau_number()}"


    def show_error(self, instance):
        box = BoxLayout(orientation='vertical', padding=10, spacing=10)
        box.add_widget(Label(text= instance))
        mybutton = Button(text='OK', size_hint_x=None, width=300, pos_hint={'center_x': 0.5})
        box.add_widget(mybutton)
        popup = Popup(title='Fehler', content=box, size_hint=(None, None), size=(500, 200))
        popup.open()
        mybutton.bind(on_press=popup.dismiss)


try:
    FolderCreator().run()
except Exception as e:
    error_message = StringIO()
    traceback.print_exc(file=error_message)
    error_message = error_message.getvalue()
    print(error_message)
    FolderCreator().show_error(error_message)